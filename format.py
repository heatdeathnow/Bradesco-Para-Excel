from openpyxl.styles import PatternFill, Side, Border, Font, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.worksheet import Worksheet
from dateutil.relativedelta import relativedelta
from openpyxl.styles.colors import Color
from openpyxl.formatting import Rule
from pandas import DataFrame
from datetime import date


def get_band(age: int) -> str:
    match age:
        case x if x <= 18:
            return '0-18'
        
        case x if x <= 23:
            return '19-23'
        
        case x if x <= 28:
            return '24-28'
        
        case x if x <= 33:
            return '29-33'
        
        case x if x <= 38:
            return '34-38'
        
        case x if x <= 43:
            return '39-43'
        
        case x if x <= 48:
            return '44-48'
        
        case x if x <= 53:
            return '49-53'

        case x if x <= 58:
            return '54-58'

        case x if x >= 59:
            return '59+'  

def clean_df(df: DataFrame) -> None:
    for i in range(len(df.index)):
        if '-' in df.iloc[i, 10]:
            df.iloc[i, 10] = '-' + df.iloc[i, 10].replace('-', '')

        df.iloc[i, 10] = float(df.iloc[i, 10].replace('.', '').replace(',', '.').strip())

        if '-' in df.iloc[i, 11]:
            df.iloc[i, 11] = '-' + df.iloc[i, 11].replace('-', '')

        df.iloc[i, 11] = float(df.iloc[i, 11].replace('.', '').replace(',', '.').strip())

        if df.iloc[i, 2] is not None and df.iloc[i, 2] != '':
            y = int(df.iloc[i, 2][6:])
            m = int(df.iloc[i, 2][3:5])
            d = int(df.iloc[i, 2][0:2])
            df.iloc[i, 2] = date(y, m, d)
        
        if df.iloc[i, 7] is not None and df.iloc[i, 7] != '':
            y = int(df.iloc[i, 7][6:])
            m = int(df.iloc[i, 7][3:5])
            d = int(df.iloc[i, 7][0:2])
            df.iloc[i, 7] = date(y, m, d)

def xlsx_format(ws: Worksheet) -> None:
    # Formatação normal
    headers = True
    header_fill = PatternFill(start_color = Color('002060'), fill_type = 'solid')
    header_font = Font(bold = True, color = Color('FFFFFF'))
    side = Side(color = Color('000000'), border_style = 'thin')
    grid = Border(side, side, side, side)

    # Formatação condicional
    green = DifferentialStyle(font = Font(bold = True), fill = PatternFill(bgColor = '228B22', fill_type = 'solid'))
    red = DifferentialStyle(font = Font(bold = True), fill = PatternFill(bgColor = 'E9967A', fill_type = 'solid'))

    last_row = ''
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = f'A2:L{ws.max_row}'
    for i, row in enumerate(ws.iter_rows(min_row = 2), 2):
        if headers:  # Para os cabeçários
            for cell in row:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = grid
            headers = False
            continue
        
        copy = False
        for j, cell in enumerate(row):
            if j == 0 and (cell.value is None or cell.value == ''):  # Vê se essa linha vai estar vazia.
                copy = True
            
            if copy and j < 6:
                cell.value = last_row[j].value  # Copia as informações da linha acima até a coluna de parentesco.
            cell.border = grid

        ws.conditional_formatting.add(f'$A{i}:$L{i}', Rule('expression', dxf = green, formula = [f'=SEARCH("I", $K{i})']))
        ws.conditional_formatting.add(f'$A{i}:$L{i}', Rule('expression', dxf = red, formula = [f'=SEARCH("C", $K{i})']))
        row[2].number_format = 'DD/MM/YYYY'
        row[7].number_format = 'DD/MM/YYYY'
        last_row = row
    
    ws['L1'] = 'Lançamento'
    ws['L1'].font = header_font
    ws['L1'].fill = header_fill
    ws['L1'].border = grid
    ws['L1'].alignment = Alignment('center', 'center')
    ws.merge_cells('L1:N1')

def add_formulaic_cols(ws: Worksheet) -> None:
    ws.insert_cols(4, 2)
    ws['D2'] = 'Idade'
    ws['E2'] = 'Faixa'

    for row in ws.iter_rows(min_row = 3):
        if row[2].value is not None and row[2].value != '':
            row[3].value = relativedelta(date.today(), row[2].value).years
            row[4].value = get_band(row[3].value)
